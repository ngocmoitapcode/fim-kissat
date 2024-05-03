import time
import os
from openpyxl import Workbook
import random
import pandas as pd
import matplotlib.pyplot as plt
import extrafunction as ef
import cnf_encoding as ce
import extrafunction as ef

def generate_input(n_items, n_transactions, output):
    with open(output, "w") as f:
        for i in range(n_transactions):
            for j in range(n_items):
                is_true = random.choice([0, 1])
                f.write(str(is_true) + " ")
            f.write("\n")

def write_data_to_excel(data, path, is_raw = True):
    if is_raw:
        # Generate header
        raw_header = data[0].keys()
        header = []
        for key in raw_header:
            header.append(key)
        # write to excel
        book = Workbook()
        sheet = book.active
        sheet.append(header)
        for d in data:
            sheet.append([d[key] for key in raw_header])
        book.save(path)
        return

    # Generate header
    raw_header = data[0].keys()
    header = []
    sub_header = []
    list_need_merge = []
    cur_col = 'A'
    header_merge = ""
    start_merge_cel = ""
    end_merge_cel = ""
    for idx, key in enumerate(raw_header):
        num_sub_header = len(key.split("/"))
        if num_sub_header > 1:
            cur_header = key.split("/")[0]
            header.append(cur_header)
            sub_header.append(key.split("/")[1])

            if header_merge != cur_header:
                if start_merge_cel != "" and end_merge_cel != "" and start_merge_cel != end_merge_cel:
                    list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
                start_merge_cel = cur_col + str(1)
                header_merge = cur_header

            end_merge_cel = cur_col + str(1)
            if idx == len(raw_header) - 1:
                list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
        else:
            header.append(key)
            sub_header.append("")
            list_need_merge.append(f'{cur_col + str(1)}:{cur_col + str(2)}')
        cur_col = chr(ord(cur_col) + 1)
    # write to excel
    book = Workbook()
    sheet = book.active
    sheet.append(header)
    sheet.append(sub_header)
    for d in data:
        sheet.append([d[key] for key in raw_header])
    for merge in list_need_merge:
        sheet.merge_cells(merge)
    book.save(path)

# def write_data_to_graph(data_from_path, output_path):
#     min_support_path = f'{output_path}/by_min_support'
#     transactions_path = f'{output_path}/by_transactions'

#     # create output path if not exist
#     os.makedirs(min_support_path, exist_ok=True)
#     os.makedirs(transactions_path, exist_ok=True)

#     # read data from excel
#     df = pd.read_excel(data_from_path)
#     for n_items in df["num_items"].unique():
#         for n_transactions in df["num_transactions"].unique():
#             # create a single figure with subplots for each combination of n_items and n_transactions
#             fig, axs = plt.subplots(1, 3, figsize=(18, 5))
#             fig.suptitle(f"Number of items: {n_items}, Number of transactions: {n_transactions}")

#             df_filtered = df[(df["num_items"] == n_items) & (df["num_transactions"] == n_transactions)]

#             # subplot 1: number of clauses
#             axs[0].plot(df_filtered["min_support_by_percent"], df_filtered["Native/clauses"], label="Native", color="red")
#             axs[0].plot(df_filtered["min_support_by_percent"], df_filtered["Nvar/clauses"], label="Nvar", color="blue")
#             axs[0].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/clauses"], label="Sq", color="yellow")
#             axs[0].set_xlabel("Minimum support")
#             axs[0].set_ylabel("Number of clauses")
#             axs[0].legend()

#             # subplot 2: time
#             axs[1].plot(df_filtered["min_support_by_percent"], df_filtered["Native/time"], label="Native", color="red")
#             axs[1].plot(df_filtered["min_support_by_percent"], df_filtered["Nvar/time"], label="Nvar", color="blue")
#             axs[1].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/time"], label="Sq", color="yellow")
#             axs[1].set_xlabel("Minimum support")
#             axs[1].set_ylabel("Time")
#             axs[1].legend()

#             # subplot 3: number of variables
#             axs[2].plot(df_filtered["min_support_by_percent"], df_filtered["Native/vars"], label="Native", color="red")
#             axs[2].plot(df_filtered["min_support_by_percent"], df_filtered["Nvar/vars"], label="Nvar", color="blue")
#             axs[2].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/vars"], label="Sq", color="yellow")
#             axs[2].set_xlabel("Minimum support")
#             axs[2].set_ylabel("Number of variables")
#             axs[2].legend()

#             plt.tight_layout(pad=3.0)
#             plt.savefig(f'{min_support_path}/n_trans_{n_transactions}.png')
#             plt.clf()

#         for min_support in df["min_support_by_percent"].unique():
#             # create a single figure with subplots for each combination of n_items and min_support
#             fig, axs = plt.subplots(1, 3, figsize=(18, 5))
#             fig.suptitle(f"Number of items: {n_items}, Minimum support: {min_support}")

#             df_filtered = df[(df["num_items"] == n_items) & (df["min_support_by_percent"] == min_support)]

#             # subplot 1: number of clauses
#             axs[0].plot(df_filtered["num_transactions"], df_filtered["Native/clauses"], label="Native", color="red")
#             axs[0].plot(df_filtered["num_transactions"], df_filtered["Nvar/clauses"], label="Nvar", color="blue")
#             axs[0].plot(df_filtered["num_transactions"], df_filtered["Sq/clauses"], label="Sq", color="yellow")
#             axs[0].set_xlabel("Number of transactions")
#             axs[0].set_ylabel("Number of clauses")
#             axs[0].legend()

#             # subplot 2: time
#             axs[1].plot(df_filtered["num_transactions"], df_filtered["Native/time"], label="Native", color="red")
#             axs[1].plot(df_filtered["num_transactions"], df_filtered["Nvar/time"], label="Nvar", color="blue")
#             axs[1].plot(df_filtered["num_transactions"], df_filtered["Sq/time"], label="Sq", color="yellow")
#             axs[1].set_xlabel("Number of transactions")
#             axs[1].set_ylabel("Time")
#             axs[1].legend()

#             # subplot 3: number of variables
#             axs[2].plot(df_filtered["num_transactions"], df_filtered["Native/vars"], label="Native", color="red")
#             axs[2].plot(df_filtered["num_transactions"], df_filtered["Nvar/vars"], label="Nvar", color="blue")
#             axs[2].plot(df_filtered["num_transactions"], df_filtered["Sq/vars"], label="Sq", color="yellow")
#             axs[2].set_xlabel("Number of transactions")
#             axs[2].set_ylabel("Number of variables")
#             axs[2].legend()

#             plt.tight_layout(pad=3.0)
#             plt.savefig(f'{transactions_path}/min_supp_{min_support}.png')
#             plt.clf()

def write_data_to_graph(data_from_path, output_path):
    min_support_path = f'{output_path}/by_min_support'
    transactions_path = f'{output_path}/by_transactions'

    # create output path if not exist
    os.makedirs(min_support_path, exist_ok=True)
    os.makedirs(transactions_path, exist_ok=True)

    # read data from excel
    df = pd.read_excel(data_from_path)
    for n_items in df["num_items"].unique():
        for n_transactions in df["num_transactions"].unique():
            # create a single figure with subplots for each combination of n_items and n_transactions
            fig, axs = plt.subplots(1, 3, figsize=(18, 5))
            fig.suptitle(f"Number of items: {n_items}, Number of transactions: {n_transactions}")

            df_filtered = df[(df["num_items"] == n_items) & (df["num_transactions"] == n_transactions)]

            # subplot 1: number of clauses
            axs[0].plot(df_filtered["min_support_by_percent"], df_filtered["Native/clauses"], label="Native", color="red")
            axs[0].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/clauses"], label="Sq", color="yellow")
            axs[0].set_xlabel("Minimum support")
            axs[0].set_ylabel("Number of clauses")
            axs[0].legend()

            # subplot 2: time
            axs[1].plot(df_filtered["min_support_by_percent"], df_filtered["Native/time"], label="Native", color="red")
            axs[1].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/time"], label="Sq", color="yellow")
            axs[1].set_xlabel("Minimum support")
            axs[1].set_ylabel("Time")
            axs[1].legend()

            # subplot 3: number of variables
            axs[2].plot(df_filtered["min_support_by_percent"], df_filtered["Native/vars"], label="Native", color="red")
            axs[2].plot(df_filtered["min_support_by_percent"], df_filtered["Sq/vars"], label="Sq", color="yellow")
            axs[2].set_xlabel("Minimum support")
            axs[2].set_ylabel("Number of variables")
            axs[2].legend()

            plt.tight_layout(pad=3.0)
            plt.savefig(f'{min_support_path}/n_trans_{n_transactions}.png')
            plt.clf()

        for min_support in df["min_support_by_percent"].unique():
            # create a single figure with subplots for each combination of n_items and min_support
            fig, axs = plt.subplots(1, 3, figsize=(18, 5))
            fig.suptitle(f"Number of items: {n_items}, Minimum support: {min_support}")

            df_filtered = df[(df["num_items"] == n_items) & (df["min_support_by_percent"] == min_support)]

            # subplot 1: number of clauses
            axs[0].plot(df_filtered["num_transactions"], df_filtered["Native/clauses"], label="Native", color="red")
            axs[0].plot(df_filtered["num_transactions"], df_filtered["Sq/clauses"], label="Sq", color="yellow")
            axs[0].set_xlabel("Number of transactions")
            axs[0].set_ylabel("Number of clauses")
            axs[0].legend()

            # subplot 2: time
            axs[1].plot(df_filtered["num_transactions"], df_filtered["Native/time"], label="Native", color="red")
            axs[1].plot(df_filtered["num_transactions"], df_filtered["Sq/time"], label="Sq", color="yellow")
            axs[1].set_xlabel("Number of transactions")
            axs[1].set_ylabel("Time")
            axs[1].legend()

            # subplot 3: number of variables
            axs[2].plot(df_filtered["num_transactions"], df_filtered["Native/vars"], label="Native", color="red")
            axs[2].plot(df_filtered["num_transactions"], df_filtered["Sq/vars"], label="Sq", color="yellow")
            axs[2].set_xlabel("Number of transactions")
            axs[2].set_ylabel("Number of variables")
            axs[2].legend()

            plt.tight_layout(pad=3.0)
            plt.savefig(f'{transactions_path}/min_supp_{min_support}.png')
            plt.clf()

# output_cnf_path is path to file .cnf
def run_multiple_input(inputfile_path, freq, use_method, output_cnf_path, all_solutions_output_folder = ''):
    TDB = ef.generate_TDB_from_input(inputfile_path)
    num_transactions, num_items = TDB.shape
    numb_vars, num_clauses = ce.generate_all_clauses(ce.all_clauses, TDB, freq, use_method, output_cnf_path)
    start_time = time.time()
    numb_solutions = ef.find_all_solutions(output_cnf_path, all_solutions_output_folder, 1, num_items, num_transactions, freq)
    total_time = time.time() - start_time
    return numb_solutions, numb_vars, num_clauses, total_time

# def benchmark():
#     os.system("bash remove.sh")
#     # generate input
#     inputs = []
#     #
#     cnf_output = []
#     #
#     all_solutions_output_folder = []
#     #
#     n_items = 5
#     for n_transactions in [20,22,24,26,28,30]:
#         for min_support_bypercent in [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]:
#             inputfile_path = f'./benchmark/input/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}.txt'
#             #
#             cnf_output.append(f'/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}.cnf')
#             #
#             path = f'/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}'
#             if not os.path.exists('./benchmark/all_solutions/native' + path):
#                 os.makedirs('./benchmark/all_solutions/native' + path)
#             if not os.path.exists('./benchmark/all_solutions/nvar' + path):
#                 os.makedirs('./benchmark/all_solutions/nvar' + path)
#             if not os.path.exists('./benchmark/all_solutions/sq' + path):
#                 os.makedirs('./benchmark/all_solutions/sq' + path)
#             all_solutions_output_folder.append(path)
#             #
#             generate_input(n_items, n_transactions, inputfile_path)
#             inputs.append((n_transactions, n_items, min_support_bypercent, inputfile_path))

#     excel_results = []

#     i = 0 # counter for the respective cnf_file_index
#     for n_transactions, n_items, min_support_bypercent, inputfile_path in inputs:
#         result = {
#             "input_path": inputfile_path,
#             "num_items": n_items,
#             "num_transactions": n_transactions,
#             "min_support_by_percent": min_support_bypercent
#         }
#         # native
#         ce.all_clauses = [] ## reset all_clauses
#         n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
#             inputfile_path = inputfile_path,
#             freq = int(min_support_bypercent * n_transactions),
#             use_method = 0,
#             output_cnf_path = './benchmark/cnf/native' + cnf_output[i],
#             all_solutions_output_folder = './benchmark/all_solutions/native' + all_solutions_output_folder[i])

#         print("Native encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
#         result["Native/vars"] = n_vars
#         result["Native/clauses"] = n_clauses
#         result["Native/solutions"] = n_solutions
#         result["Native/time"] = elapsed_time

#         # nvar
#         ce.all_clauses = [] ## reset all_clauses
#         n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
#             inputfile_path = inputfile_path,
#             freq = int(min_support_bypercent * n_transactions),
#             use_method = 1,
#             output_cnf_path = './benchmark/cnf/nvar' + cnf_output[i],
#             all_solutions_output_folder = './benchmark/all_solutions/nvar' + all_solutions_output_folder[i])
#         print("Nvar encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
#         result["Nvar/vars"] = n_vars
#         result["Nvar/clauses"] = n_clauses
#         result["Nvar/solutions"] = n_solutions
#         result["Nvar/time"] = elapsed_time

#         # sq
#         ce.all_clauses = [] ## reset all_clauses
#         n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
#             inputfile_path = inputfile_path,
#             freq = int(min_support_bypercent * n_transactions),
#             use_method = 2,
#             output_cnf_path = './benchmark/cnf/sq' + cnf_output[i],
#             all_solutions_output_folder = './benchmark/all_solutions/sq' + all_solutions_output_folder[i])
#         print("Sq encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
#         result["Sq/vars"] = n_vars
#         result["Sq/clauses"] = n_clauses
#         result["Sq/solutions"] = n_solutions
#         result["Sq/time"] = elapsed_time

#         excel_results.append(result)
#         i += 1
#     #get unique file name
#     t = time.strftime("%Y%m%d_%H%m")
#     # create output path if not exist
#     output_path = f'./benchmark/excel/{t}'
#     os.makedirs(output_path, exist_ok=True)
#     raw_output_path = f'./benchmark/excel/{t}/benchmark_raw.xlsx'
#     beauty_output_path = f'./benchmark/excel/{t}/benchmark.xlsx'
#     write_data_to_excel(excel_results, raw_output_path)
#     write_data_to_excel(excel_results, beauty_output_path, False)

#     write_data_to_graph(raw_output_path, output_path)

def benchmark():
    os.system("bash remove.sh")
    # generate input
    inputs = []
    #
    cnf_output = []
    #
    all_solutions_output_folder = []
    #
    n_items = 5
    for n_transactions in [20,21,22,23,24,25,26]:
        for min_support_bypercent in [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]:
            inputfile_path = f'./benchmark/input/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}.txt'
            #
            cnf_output.append(f'/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}.cnf')
            #
            path = f'/{n_items}_items_{n_transactions}_trans_{min_support_bypercent}'
            if not os.path.exists('./benchmark/all_solutions/native' + path):
                os.makedirs('./benchmark/all_solutions/native' + path)
            if not os.path.exists('./benchmark/all_solutions/sq' + path):
                os.makedirs('./benchmark/all_solutions/sq' + path)
            all_solutions_output_folder.append(path)
            #
            generate_input(n_items, n_transactions, inputfile_path)
            inputs.append((n_transactions, n_items, min_support_bypercent, inputfile_path))

    excel_results = []

    i = 0 # counter for the respective cnf_file_index
    for n_transactions, n_items, min_support_bypercent, inputfile_path in inputs:
        result = {
            "input_path": inputfile_path,
            "num_items": n_items,
            "num_transactions": n_transactions,
            "min_support_by_percent": min_support_bypercent
        }
        # native
        ce.all_clauses = [] ## reset all_clauses
        n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
            inputfile_path = inputfile_path,
            freq = int(min_support_bypercent * n_transactions),
            use_method = 0,
            output_cnf_path = './benchmark/cnf/native' + cnf_output[i],
            all_solutions_output_folder = './benchmark/all_solutions/native' + all_solutions_output_folder[i])

        print("Native encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
        result["Native/vars"] = n_vars
        result["Native/clauses"] = n_clauses
        result["Native/solutions"] = n_solutions
        result["Native/time"] = elapsed_time

        # sq
        ce.all_clauses = [] ## reset all_clauses
        n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
            inputfile_path = inputfile_path,
            freq = int(min_support_bypercent * n_transactions),
            use_method = 2,
            output_cnf_path = './benchmark/cnf/sq' + cnf_output[i],
            all_solutions_output_folder = './benchmark/all_solutions/sq' + all_solutions_output_folder[i])
        print("Sq encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
        result["Sq/vars"] = n_vars
        result["Sq/clauses"] = n_clauses
        result["Sq/solutions"] = n_solutions
        result["Sq/time"] = elapsed_time

        excel_results.append(result)
        i += 1
    #get unique file name
    t = time.strftime("%Y%m%d_%H%m")
    # create output path if not exist
    output_path = f'./benchmark/excel/{t}'
    os.makedirs(output_path, exist_ok=True)
    raw_output_path = f'./benchmark/excel/{t}/benchmark_raw.xlsx'
    beauty_output_path = f'./benchmark/excel/{t}/benchmark.xlsx'
    write_data_to_excel(excel_results, raw_output_path)
    write_data_to_excel(excel_results, beauty_output_path, False)

    write_data_to_graph(raw_output_path, output_path)

    
    
# if __name__ == "__main__":
#     benchmark()
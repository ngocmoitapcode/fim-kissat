import time
import os
from openpyxl import Workbook
import random
import pandas as pd
import matplotlib.pyplot as plt
import extrafunction as ef
import cnf_encoding as ce
import extrafunction as ef
import glob
import numpy as np

from cpmpy import *
from cpmpy.solvers import *

def write_data_to_excel(data, path, is_raw = True):
    if is_raw:
        # Generate header
        raw_header = data[0].keys()
        header = []
        for key in raw_header:
            header.append(key)
        # write to excel
        book = Workbook()
        sheet = book.active
        sheet.append(header)
        for d in data:
            sheet.append([d[key] for key in raw_header])
        book.save(path)
        return

    # Generate header
    raw_header = data[0].keys()
    header = []
    sub_header = []
    list_need_merge = []
    cur_col = 'A'
    header_merge = ""
    start_merge_cel = ""
    end_merge_cel = ""
    for idx, key in enumerate(raw_header):
        num_sub_header = len(key.split("/"))
        if num_sub_header > 1:
            cur_header = key.split("/")[0]
            header.append(cur_header)
            sub_header.append(key.split("/")[1])

            if header_merge != cur_header:
                if start_merge_cel != "" and end_merge_cel != "" and start_merge_cel != end_merge_cel:
                    list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
                start_merge_cel = cur_col + str(1)
                header_merge = cur_header

            end_merge_cel = cur_col + str(1)
            if idx == len(raw_header) - 1:
                list_need_merge.append(f'{start_merge_cel}:{end_merge_cel}')
        else:
            header.append(key)
            sub_header.append("")
            list_need_merge.append(f'{cur_col + str(1)}:{cur_col + str(2)}')
        cur_col = chr(ord(cur_col) + 1)
    # write to excel
    book = Workbook()
    sheet = book.active
    sheet.append(header)
    sheet.append(sub_header)
    for d in data:
        sheet.append([d[key] for key in raw_header])
    for merge in list_need_merge:
        sheet.merge_cells(merge)
    book.save(path)

def write_data_to_graph(data_from_path, output_path):
    df = pd.read_excel(data_from_path)

    # List of datasets
    datasets = ['lymph', 'tictactoe', 'zoo1', 'vote']

    # Loop over the datasets
    for dataset in datasets:
        # Filter the data for the current dataset
        df_filtered = df[df['dataset'] == dataset]

        # Create a new figure
        fig, ax = plt.subplots(figsize=(6, 5))

        # Plot the data for the two methods
        ax.plot(df_filtered['min_support_by_percent'], df_filtered['Cpm/time'], label='cpmpy', color='red')
        ax.plot(df_filtered['min_support_by_percent'], df_filtered['Sq/time'], label='sq', color='blue')

        # Set the labels and title
        ax.set_xlabel('Minimum support')
        ax.set_ylabel('Time')
        ax.set_title(f'Dataset: {dataset}')

        # Add a legend
        ax.legend()

        # Save the figure
        plt.savefig(f'{output_path}/{dataset}.png')

        # Close the figure
        plt.close(fig)

# output_cnf_path is path to file .cnf
def run_multiple_input(inputfile_path, freq, use_method, output_cnf_path, all_solutions_output_folder = ''):
    TDB = ef.generate_TDB_from_input(inputfile_path)
    num_transactions, num_items = TDB.shape
    numb_vars, num_clauses = ce.generate_all_clauses(ce.all_clauses, TDB, freq, use_method, output_cnf_path)
    start_time = time.time()
    numb_solutions = ef.find_all_solutions(output_cnf_path, all_solutions_output_folder, 1, num_items, num_transactions, freq)
    total_time = time.time() - start_time
    return numb_solutions, numb_vars, num_clauses, total_time

def cpm_solve_fim(input_file, freq):
    # Load the data
    TDB = []
    with open(input_file) as f:
        for line in f:
            TDB.append([bool(int(x)) for x in line.split()])
    f.close()
    TDB = np.array(TDB)
    nrT, nrI = TDB.shape

    Items = boolvar(shape=nrI, name="items")
    Trans = boolvar(shape=nrT, name="transactions")

    m = Model()
    for T, negrow in zip(Trans, ~TDB):
        m += T == ~any(Items[np.where(negrow)])
    m += (sum(Trans) >= freq)
    # Solve the problem
    start_time = time.time()
    n_solutions = m.solveAll(solver='pysat')
    elapsed_time = time.time() - start_time

    return n_solutions, elapsed_time

def benchmarkR():
    os.system("bash removeR.sh")
    # generate input
    inputs = []
    #
    cnf_output = []
    #
    all_solutions_output_folder = []
    #
    # Get all .txt files in the dataset-convert folder
    input_files = glob.glob('./dataset-convert/*.txt')

    for inputfile_path in input_files:
        # Load the data and get its shape
        TDB = np.loadtxt(inputfile_path)
        n_transactions, n_items = TDB.shape
        # Extract the filename from the path
        filename = os.path.basename(inputfile_path)

        for min_support_bypercent in [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]:
            #
            cnf_output.append(f'/{filename}_{n_items}_items_{n_transactions}_trans_{min_support_bypercent}.cnf')
            #
            path = f'/{filename}_{n_items}_items_{n_transactions}_trans_{min_support_bypercent}'
            if not os.path.exists('./benchmarkR/all_solutions/sq' + path):
                os.makedirs('./benchmarkR/all_solutions/sq' + path)
            all_solutions_output_folder.append(path)
            #
            inputs.append((n_transactions, n_items, min_support_bypercent, inputfile_path))

    excel_results = []

    i = 0 # counter for the respective cnf_file_index
    for n_transactions, n_items, min_support_bypercent, inputfile_path in inputs:
        result = {
            "input_path": inputfile_path,
            "num_items": n_items,
            "num_transactions": n_transactions,
            "min_support_by_percent": min_support_bypercent
        }
        # cpm
        n_solutions, elapsed_time = cpm_solve_fim(
            input_file = inputfile_path,
            freq = int(min_support_bypercent * n_transactions))

        print("Cpm encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions)
        result["Cpm/solutions"] = n_solutions
        result["Cpm/time"] = elapsed_time

        # sq
        ce.all_clauses = [] ## reset all_clauses
        n_solutions, n_vars, n_clauses, elapsed_time = run_multiple_input(
            inputfile_path = inputfile_path,
            freq = int(min_support_bypercent * n_transactions),
            use_method = 2,
            output_cnf_path = './benchmarkR/cnf/sq' + cnf_output[i],
            all_solutions_output_folder = './benchmarkR/all_solutions/sq' + all_solutions_output_folder[i])
        print("Sq encoding:", n_items, n_transactions, min_support_bypercent, int(min_support_bypercent * n_transactions), n_solutions, n_vars, n_clauses)
        print(cnf_output[i])
        result["Sq/solutions"] = n_solutions
        result["Sq/time"] = elapsed_time

        excel_results.append(result)
        i += 1
    #get unique file name
    t = time.strftime("%Y%m%d_%H%m")
    # create output path if not exist
    output_path = f'./benchmarkR/excel/{t}'
    os.makedirs(output_path, exist_ok=True)
    raw_output_path = f'./benchmarkR/excel/{t}/benchmarkR_raw.xlsx'
    beauty_output_path = f'./benchmarkR/excel/{t}/benchmarkR.xlsx'
    write_data_to_excel(excel_results, raw_output_path)
    write_data_to_excel(excel_results, beauty_output_path, False)

    write_data_to_graph(raw_output_path, output_path)

# if __name__ == "__main__":
#     benchmarkR()